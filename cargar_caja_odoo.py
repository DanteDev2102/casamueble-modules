#!/usr/bin/env python3
"""
cargar_caja_odoo.py - v3
-----------------------
Escaneo dinámico de hojas de Excel para cargar Ingresos y Egresos en Odoo.
Busca cualquier hoja con 'INGRESOS' o 'EGRESOS' en el nombre.
"""

import os
import openpyxl
from datetime import datetime, date
from pathlib import Path
import xmlrpc.client

# ────────────────────────────────────────────────────────────────
#  CONFIG
# ────────────────────────────────────────────────────────────────
ODOO_URL = "http://192.168.10.103:17000"
ODOO_DB = "casamueble"
ODOO_USER = "casamuebleca@gmail.com"
ODOO_PASSWORD = "8bc0615d3eb19925ef3607965de0ab6080fa47c9"

BASE_DIR = Path(__file__).parent

# Archivos a procesar
ARCHIVOS_POSIBLES = [
    "enero.xlsx",
    "febrero.xlsx",
    "CAJA ENERO 2026.xlsx",
    "CAJA FEBRERO 2026.xlsx",
]

# Diarios contables (unique codes)
DIARIOS_CONFIG = {
    "bs": {"name": "Efectivo VES", "code": "EFVES", "tipo": "cash"},
    "zelle": {"name": "Zelle", "code": "ZELLE", "tipo": "bank"},
    "efectivo": {"name": "Efectivo USD", "code": "EFUSD", "tipo": "cash"},
    "panama": {"name": "Panamá", "code": "PANAM", "tipo": "bank"},
}

# Clasificación de TIPOS para facturas
TIPOS_VENTA = {"CXC", "VENTAS NUEVAS", "VENTAS CONTADO", "VENTAS NUEVA"}
TIPOS_PAGADA = {"VENTAS CONTADO"}
TIPOS_COMPRA = {
    "MATERIA PRIMA",
    "GASTO VARIABLE",
    "MATERIALES",
    "INTERNET TIENDA",
    "INTERNET TALLER",
    "NOMINA TIENDA",
    "NOMINA TALLER",
    "REMODELACIÓN",
    "ALQUILER TIENDA",
    "NÓMINA TIENDA",
    "NÓMINA TALLER",
    "PRODUCCION TALL",
    "PRODUCCION TALLER",
    "PRODUCCIÓN TALLER",
}

# ────────────────────────────────────────────────────────────────
#  XML-RPC & HELPERS
# ────────────────────────────────────────────────────────────────


def conectar():
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common", allow_none=True)
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PASSWORD, {})
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object", allow_none=True)
    return uid, models


def ex(models, uid, model, method, *args, **kw):
    return models.execute_kw(ODOO_DB, uid, ODOO_PASSWORD, model, method, list(args), kw)


def ex_safe(models, uid, model, method, *args, **kw):
    try:
        return ex(models, uid, model, method, *args, **kw)
    except:
        return None


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace("$", "").replace(" ", ""))
    except:
        return 0.0


def fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).strftime("%Y-%m-%d")
            except:
                pass
    return datetime.today().strftime("%Y-%m-%d")


# ────────────────────────────────────────────────────────────────
#  LECTURA DE EXCEL (DINÁMICA)
# ────────────────────────────────────────────────────────────────


def _detectar_cols(row):
    mapping = {}
    for i, cell in enumerate(row):
        if cell:
            mapping[str(cell).strip().upper().rstrip(".")] = i
    return mapping


ALIAS = {
    "FECHA": ["FECHA"],
    "TIPO": ["TIPO"],
    "DESC": ["DESCRIPCIÓN", "DESCRIPCION", "DESCRIPCI", "DESCRIP"],
    "BS": ["BS", "BS.", "BOLIVARES"],
    "ZELLE": ["ZELLE"],
    "EFECTIVO": ["EFECTIVO"],
    "PANAMA": ["PANAMÁ", "PANAMA"],
    "DETALLES": ["DETALLES", "DETALLE", "OBSERVACIONES"],
}


def _parse_row(row, col_map):
    def g(campo):
        for a in ALIAS.get(campo, [campo]):
            if a in col_map and col_map[a] < len(row):
                return row[col_map[a]]
        return None

    fecha = fmt_date(g("FECHA"))
    tipo = str(g("TIPO") or "").strip().upper()
    desc = str(g("DESC") or "").strip() or str(g("DETALLES") or "").strip()

    bs, zelle, efectivo, panama = (
        safe_float(g("BS")),
        safe_float(g("ZELLE")),
        safe_float(g("EFECTIVO")),
        safe_float(g("PANAMA")),
    )
    total = bs + zelle + efectivo + panama

    if total == 0 or not tipo or tipo == "TIPO":
        return None
    return {
        "fecha": fecha,
        "tipo": tipo,
        "desc": desc,
        "bs": bs,
        "zelle": zelle,
        "efectivo": efectivo,
        "panama": panama,
        "total": total,
        "detalles": str(g("DETALLES") or ""),
        "ref": f"{fecha}|{tipo}|{desc[:30]}|{total:.2f}",
    }


def _es_encabezado(row):
    vals = [str(v).strip().upper() for v in row if v]
    return "FECHA" in vals and "TIPO" in vals


def _leer_hoja(ws):
    items, col_map = [], {}
    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        if _es_encabezado(row):
            col_map = _detectar_cols(row)
            continue
        if col_map:
            p = _parse_row(row, col_map)
            if p:
                items.append(p)
    return items


def leer_secciones(path: Path):
    print(f"   🔍 Leyendo '{path.name}'...")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ingresos, egresos = [], []
    for sn in wb.sheetnames:
        upper = sn.upper()
        if "INGRESOS" in upper:
            print(f"      📍 Escaneando Ingresos: {sn}")
            ingresos.extend(_leer_hoja(wb[sn]))
        elif "EGRESOS" in upper:
            print(f"      📍 Escaneando Egresos: {sn}")
            egresos.extend(_leer_hoja(wb[sn]))
    wb.close()
    return ingresos, egresos


# ────────────────────────────────────────────────────────────────
#  ODOO OPS
# ────────────────────────────────────────────────────────────────


def buscar_cuenta(models, uid, palabras, fallback="income"):
    for p in palabras:
        ids = ex(
            models,
            uid,
            "account.account",
            "search",
            [("name", "ilike", p), ("deprecated", "=", False)],
        )
        if ids:
            return ids[0]
    ids = ex(
        models,
        uid,
        "account.account",
        "search",
        [("account_type", "=", fallback), ("deprecated", "=", False)],
    )
    return ids[0] if ids else False


def asegurar_partner(models, uid, nombre, prov=False, cli=False):
    nombre = (nombre or "Varios").strip()
    ids = ex(models, uid, "res.partner", "search", [("name", "=ilike", nombre)])
    if ids:
        return ids[0]
    return ex(
        models,
        uid,
        "res.partner",
        "create",
        {
            "name": nombre,
            "supplier_rank": 1 if prov else 0,
            "customer_rank": 1 if cli else 0,
        },
    )


def asegurar_diario(models, uid, name, code, tipo="cash"):
    ids = ex(models, uid, "account.journal", "search", [("name", "=ilike", name)])
    if ids:
        return ids[0]
    # Buscar por código por si acaso
    ids = ex(models, uid, "account.journal", "search", [("code", "=", code)])
    if ids:
        return ids[0]
    return ex(
        models,
        uid,
        "account.journal",
        "create",
        {"name": name, "type": tipo, "code": code},
    )


def crear_factura(models, uid, row, es_ing):
    ref = f"{'ING' if es_ing else 'EGR'}|{row['ref']}"
    if ex(models, uid, "account.move", "search", [("ref", "=", ref)]):
        return None, False
    pid = asegurar_partner(models, uid, row["desc"], prov=not es_ing, cli=es_ing)
    c_id = buscar_cuenta(
        models,
        uid,
        ["Ventas"] if es_ing else ["Compras", "Gasto"],
        "income" if es_ing else "expense",
    )
    mid = ex(
        models,
        uid,
        "account.move",
        "create",
        {
            "move_type": "out_invoice" if es_ing else "in_invoice",
            "partner_id": pid,
            "invoice_date": row["fecha"],
            "ref": ref,
            "invoice_line_ids": [
                (
                    0,
                    0,
                    {
                        "name": row["detalles"] or row["tipo"],
                        "price_unit": row["total"],
                        "quantity": 1.0,
                        "account_id": c_id,
                    },
                )
            ],
        },
    )
    ex_safe(models, uid, "account.move", "action_post", [mid])
    return mid, True


def crear_asiento(models, uid, row, es_ing):
    ref = f"{'ING' if es_ing else 'EGR'}|{row['ref']}"
    if ex(models, uid, "account.move", "search", [("ref", "=", ref)]):
        return None, False
    mid = ex(
        models,
        uid,
        "account.move",
        "create",
        {
            "move_type": "entry",
            "date": row["fecha"],
            "ref": ref,
            "line_ids": [
                (
                    0,
                    0,
                    {
                        "name": row["tipo"],
                        "account_id": buscar_cuenta(
                            models, uid, ["Caja", "Banco"], "asset_receivable"
                        ),
                        "debit": row["total"] if es_ing else 0,
                        "credit": 0 if es_ing else row["total"],
                    },
                ),
                (
                    0,
                    0,
                    {
                        "name": row["tipo"],
                        "account_id": buscar_cuenta(
                            models,
                            uid,
                            ["Ingreso"] if es_ing else ["Gasto"],
                            "income" if es_ing else "expense",
                        ),
                        "debit": 0 if es_ing else row["total"],
                        "credit": row["total"] if es_ing else 0,
                    },
                ),
            ],
        },
    )
    ex_safe(models, uid, "account.move", "action_post", [mid])
    return mid, True


# ────────────────────────────────────────────────────────────────
#  MAIN
# ────────────────────────────────────────────────────────────────


def main():
    print("🚀 Cargador de Caja v3 (Escaneo total)")
    uid, models = conectar()
    # Asegurar diarios
    for k, v in DIARIOS_CONFIG.items():
        asegurar_diario(models, uid, v["name"], v["code"], v["tipo"])

    for f in ARCHIVOS_POSIBLES:
        path = BASE_DIR / f
        if not path.exists():
            continue
        print(f"\n📂 Archivo: {f}")
        ing, egr = leer_secciones(path)
        for filas, es_ing in [(ing, True), (egr, False)]:
            print(f"   ► {'Ingresos' if es_ing else 'Egresos'}: {len(filas)} filas")
            for row in filas:
                try:
                    if (es_ing and row["tipo"] in TIPOS_VENTA) or (
                        not es_ing and row["tipo"] in TIPOS_COMPRA
                    ):
                        _, ok = crear_factura(models, uid, row, es_ing)
                    else:
                        _, ok = crear_asiento(models, uid, row, es_ing)
                    if ok:
                        print(
                            f"    ✅ {row['fecha']} {row['tipo'][:15]:<15} {row['total']:>8.2f} - {row['desc'][:30]}"
                        )
                except Exception as e:
                    print(f"    ❌ Error: {e}")


if __name__ == "__main__":
    main()
